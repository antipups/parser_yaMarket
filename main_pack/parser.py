import datetime
import os
from random import randint as ri
import time
import requests
import openpyxl
from main_pack import parser_proxy
import re
from PIL import Image


def read_xlsx(path):
    """
        Открываем и считываем файл xlsx для получения всех ссылок и названий товаров
    :return:
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['Лист1']
    i = 1
    dict_of_product = {}
    while True:
        if ws.cell(row=i, column=2).value is not None:
            key = ws.cell(row=i, column=1).value
            if ws.cell(row=i, column=2).hyperlink is not None and ws.cell(row=i, column=2).hyperlink.target is not None:
                value = ws.cell(row=i, column=2).hyperlink.target
            else:
                value = ws.cell(row=i, column=2).value
            dict_of_product.update({key: value})
            i += 1
        else:
            wb.close()
            parse(dict_of_product)
            return


def parse(dict_of_product):
    """
        Парсим ВСЕ товары с Я Маркета, с использованием прокси.
    :param dict_of_product:
    :return:
    """
    # with open('asd.txt', 'r') as f:
    all_proxies = parser_proxy.parse_proxy()
        # all_proxies = f.read().split('\n')
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Cookie': '_ym_uid=1579538585539137365; mda=0; yandexuid=4775578911579527928; yuidss=4775578911579527928; i=7fpXxLZtVlknYALqfuHdIqESJ7QKPyrT/p1Ib1NKFr2+GlI6yQ4Nq8EuI7FxxNSTvC5JPW7DUgpjBnX2NCjgeBJhLQE=; my=YwA=; yandex_gid=142; gdpr=0; yandexmarket=48%2CUAH%2C1%2C%2C%2C%2C2%2C0%2C0%2C0%2C0%2C0%2C12%2C0; mOC=1; currentRegionId=142; currentRegionName=%D0%94%D0%BE%D0%BD%D0%B5%D1%86%D0%BA; yabs-frequency=/4/0000000000000000/AKgmSB0wGG00/; oMaSefD=1; oMaSpfD=1; oMaRefD=1; Session_id=3:1587567800.5.0.1587567800047:uOvH2Q:4c.1|230517873.0.2|4:156007.208530.Ct9WborQwK5ZqYGCgMhyMgHKkSI; sessionid2=3:1587567800.5.0.1587567800047:uOvH2Q:4c.1|230517873.0.2|4:156007.666179.wzxl5em2LjckToAqLylJxzwXt3M; L=WnRGfUNYc2xfWkAFakRaBg5bYgN0cAVxGAcNIUANFBMvBwIAJw==.1587567800.14211.321094.42a247e10a7e1c492bb9a0954f71a90a; yandex_login=nick.kurkurin; skid=7141896761587577073; _ym_d=1587577075; oMaSofD=1; oMaFifD=1; oMaPrfD=1; _ym_isad=2; _ym_wasSynced=%7B%22time%22%3A1588195331685%2C%22params%22%3A%7B%22eu%22%3A0%7D%2C%22bkParams%22%3A%7B%7D%7D; yp=1581234933.com.3#1581437288.dq.1#1896019512.multib.1#1584131668.oyu.4775578911579527928#1596427321.szm.1_25%3A1536x864%3A1536x760#1902927800.udn.cDpuaWNrLmt1cmt1cmlu#1612368485.ygu.1#1894898590.yrtsi.1579538590#1588281734.yu.4775578911579527928; ymex=1590787334.oyu.4775578911579527928#1898004678.yrts.1582644678#1894898590.yrtsi.1579538590; _ym_visorc_10630330=b; _ym_metrika_enabled_10630330=1; spravka=dD0xNTg4MTk1MzQ1O2k9MjE3LjE5OS4yMzYuMTUxO3U9MTU4ODE5NTM0NTUyNjA3Mzk0MTtoPTlmOTRiOTA5ZGQ5ZjUxY2U5YWY2NzAxMmJmZWE5YTRh; visits=1586198458-1587974766-1588195345; lr=0; parent_reqid_seq=7634b215cd9e2d690e76b247a5daee8a; uid=AABcfl6p8BGq2wBoBp5QAg==; js=1; dcm=1; metrika_enabled=1; _ym_metrika_enabled=1; _ym_visorc_160656=b; first_visit_time=2020-04-30T00%3A22%3A26%2B03%3A00; HISTORY_AUTH_SESSION=880ba9a1; fonts-loaded=1; _ym_visorc_45411513=b; ugcp=1; _ym_metrika_enabled_160656=1; _ym_metrika_enabled_44910898=1; _ym_metrika_enabled_45411513=1',
        'Host': 'market.yandex.ua',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36',
    }
    st = datetime.datetime.now()
    i = 0
    for index, product in enumerate(dict_of_product.items()):
        if i == 5:
            all_proxies = parser_proxy.parse_proxy()
        else:
            i += 1
        html, title, url = '', product[0].strip(), product[1][:product[1].find('?')] + ('/spec' if product[1].find('/spec') == -1 else '')
        print(f"Процесс - {index + 1}/{len(dict_of_product.keys())}", title, url)
        start = datetime.datetime.now()
        while html.find(title) == -1:
            proxies = {'https': all_proxies[ri(0, len(all_proxies) - 1)]}
            print(proxies)
            try:
                html = requests.get(url, headers=headers, proxies=proxies)
                if not html.ok:
                    print(f'Мертвая ссылка - {url}, {title}')
                    break
                else:
                    html = html.text
            except (requests.exceptions.ProxyError, requests.exceptions.SSLError):
                print('Ошибка из-за неверного прокси, продолжаю работу;')
                time.sleep(3)
                continue
            else:
                with open('temp.txt', 'w', encoding='utf-8') as f:
                    f.write(html)
                print('Времени прошло - ', datetime.datetime.now() - start)
                if get_info(title, url, html) is False:
                    continue
                break
    print('Закончил работу за - ', datetime.datetime.now() - st)


def parse_all_pic(html):
    pic_code = re.search(r'{\"property\":\"og:image\",\"content\":\".*twitter:image', html)
    if not pic_code:
        return False
    if not os.path.exists('image'):
        os.mkdir('image')

    pic_code = pic_code.group()
    all_pic = []
    all_name_of_pic = []
    for index, pic in enumerate(re.finditer(r'https((?!hq).)*hq', pic_code), start=1):
        pic = pic.group()
        if pic not in all_pic:
            all_pic.append(pic)
            filename = 'image\\' + datetime.datetime.now().strftime('%Y%m%d%H%M%S%f') + '.png'
            all_name_of_pic.append(filename)
            with open(filename, 'wb') as f:
                f.write(requests.get(pic).content)
    return tuple(all_name_of_pic)


def get_info(title, url, html):
    """
        Получаем всю инфу о продукте.
    :return:
    """
    print('Достаю инфу с спарсенного сайта.')
    # with open('Дамп.txt', 'w', encoding='utf-8') as f:
    #     f.write(html)
    tuple_of_pic = parse_all_pic(html)
    if not tuple_of_pic:
        if html.find('captcha') > -1:
            print('Ошибка парсинга, пробую по новой.')
        return False

    dict_of_info = {'Название': title, 'Ссылка': url}
    for match in re.finditer(r"_2TxqAVjiup((?!</dd).)*", html, re.MULTILINE):
        characteristic = match.group()
        key = re.search(r'<span.[^/]*', characteristic).group()[6:-1]
        value = characteristic[characteristic.rfind('>') + 1:]
        dict_of_info.update({key: value})

    write_xlsx(dict_of_info, tuple_of_pic)


def write_xlsx(dict_of_info, tuple_of_pic):
    try:
        wb = openpyxl.load_workbook('2.xlsx')
    except:
        wb = openpyxl.Workbook()
        del wb['Sheet']
    finally:
        if 'Лист1' not in wb.sheetnames:
            wb.create_sheet('Лист1')
        ws = wb['Лист1']
        ws.cell(row=2, column=2).value = 'Путь к картинке'
        ws.column_dimensions['a'].width = 30
        i = 2
        dict_of_col = {}
        while ws.cell(row=2, column=i).value is not None:
            dict_of_col.update({ws.cell(row=2, column=i).value: i})
            i += 1
        else:
            for j in dict_of_info.keys():
                if j not in dict_of_col:
                    ws.cell(row=2, column=i).value = j
                    dict_of_col.update({j:i})
                    i += 1

    print('Открыл 2.xlsx.')

    row = ws.max_row + 1 if ws.max_row == 2 else ws.max_row + 14
    for index, pic in enumerate(tuple_of_pic):
        with Image.open(pic) as img:
            img.thumbnail((200, 200), Image.ANTIALIAS)
            img.save(pic)
        img = openpyxl.drawing.image.Image(pic)
        ws.add_image(img, 'A' + str(row + 11 * index))
        ws.cell(column=2, row=row + 1 + 11 * index).value = pic
    else:
        ws.cell(row=row + 1 + 11 * index, column=1).value = 'Максимальная строка'

    for column, val in dict_of_info.items():
        ws.cell(row=row + 1, column=dict_of_col.get(column)).value = val

    while True:
        try:
          wb.save('2.xlsx')
        except Exception as e:
            print('Ошибка - ', e)
            print('Закройте пожалуйста Excel файл 2.xlsx')
            time.sleep(3)
        else:
            wb.close()
            print('Закрыл 2.xlsx.')
            return


if __name__ == '__main__':
    # get_info('sex', 'asd', 'lol')
    # write_xlsx("sex")
    # read_xlsx('../1.xlsx')
    # get_info('1', '1', '1')
    # write_xlsx(1)
    with open('Дамп.txt', 'r', encoding='utf-8') as f:
        get_info('asd', 'asd', f.read())
    # write_xlsx('sex', 25)

    input('Парсер завершил свою работу, для выхода нажмите какую-либо кнопку.')