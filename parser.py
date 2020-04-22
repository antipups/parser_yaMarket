import datetime
import random
import time
import requests
import openpyxl
import parser_proxy
import re
from PIL import Image


def read_xlsx():
    """
        Открываем и считываем файл xlsx для получения всех ссылок и названий товаров
    :return:
    """
    print('Открыл 1.xlsx')
    wb = openpyxl.load_workbook('1.xlsx')
    ws = wb['Лист1']
    i = 1
    dict_of_product = {}
    while True:
        if ws.cell(row=i, column=2).value is not None:
            dict_of_product.update({ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value})
            i += 1
        else:
            wb.close()
            break
    print('Закрыл 1.xlsx')
    parse(dict_of_product)


def parse(dict_of_product):
    """
        Парсим ВСЕ товары с Я Маркета, с использованием прокси.
    :param dict_of_product:
    :return:
    """
    all_proxies = parser_proxy.parse_proxy()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36',
        'Cookie': '_ym_uid=1579538585539137365; mda=0; yandexuid=4775578911579527928; yuidss=4775578911579527928; i=7fpXxLZtVlknYALqfuHdIqESJ7QKPyrT/p1Ib1NKFr2+GlI6yQ4Nq8EuI7FxxNSTvC5JPW7DUgpjBnX2NCjgeBJhLQE=; my=YwA=; L=XVB2anxpXG1iRl9SAVJtfm9zXWZ2Wg52JSwMOWccICc7PgpZFA==.1580660121.14129.317553.9f500dbfa9f5bde8515aefe3689df390; yandex_gid=142; gdpr=0; yandexmarket=48%2CUAH%2C1%2C%2C%2C%2C2%2C0%2C0%2C0%2C0%2C0%2C12%2C0; mOC=1; currentRegionId=142; currentRegionName=%D0%94%D0%BE%D0%BD%D0%B5%D1%86%D0%BA; cycada=vBIVPiL1OivB3zUy7Bun86lEsDT+JL2RpLjQnNHDsg4=; _ym_d=1586853678; yabs-frequency=/4/0000000000000000/AKgmSB0wGG00/; skid=6518588401587125766; pof=%7B%22clid%22%3A%5B%222359538%22%5D%2C%22mclid%22%3A%221002%22%2C%22distr_type%22%3A%227%22%2C%22vid%22%3A%22986%22%2C%22opp%22%3A%22900%22%7D; cpa-pof=%7B%22clid%22%3A%5B%222359538%22%5D%2C%22mclid%22%3A%221002%22%2C%22distr_type%22%3A%227%22%2C%22vid%22%3A%22986%22%2C%22opp%22%3A%22900%22%7D; visits=1586198458-1587125766-1587549787; uid=AABcEl6gFluZxwC0A8FCAg==; js=1; dcm=1; _ym_wasSynced=%7B%22time%22%3A1587549786030%2C%22params%22%3A%7B%22eu%22%3A0%7D%2C%22bkParams%22%3A%7B%7D%7D; first_visit_time=2020-04-22T13%3A03%3A06%2B03%3A00; fonts-loaded=1; _ym_isad=2; ymex=1590141790.oyu.4775578911579527928#1898004678.yrts.1582644678#1894898590.yrtsi.1579538590; ugcp=1; oMaSefD=1; yp=1581234933.com.3#1581437288.dq.1#1896019512.multib.1#1584131668.oyu.4775578911579527928#1596427321.szm.1_25%3A1536x864%3A1536x760#1612368485.ygu.1#1894898590.yrtsi.1579538590#1587636190.yu.4775578911579527928; ys=; cmp-merge=true; reviews-merge=true; HISTORY_UNAUTH_SESSION=true; lr=0; oMaSpfD=1; oMaRefD=1; _ym_visorc_160656=b; _ym_visorc_45411513=b; parent_reqid_seq=d499a88d668b922c6d2922077f6b8f12%2Ccbd6e7cb41bc5d9511d0dfee417e46a7%2Ccf5f49116811a02fc7fa555667a792aa%2C67e994d7a1532a062e3f3fbf892f1777%2C36e7a5208ea5ea6adbb11512a45e181c; metrika_enabled=1; _ym_metrika_enabled=1; _ym_metrika_enabled_160656=1; _ym_metrika_enabled_44910898=1; _ym_metrika_enabled_45411513=1'
    }
    st = datetime.datetime.now()
    for title, url in dict_of_product.items():
        html = ''
        title = title.strip()
        print(title, url)
        i = 0
        url = url[:url.find('?')] + '/spec'
        start = datetime.datetime.now()
        while html.find(title) == -1:
            proxies = {'https': all_proxies[random.randint(0, len(all_proxies) - 1)]}
            try:
                html = requests.get(url, headers=headers, proxies=proxies).text
            except requests.exceptions.ProxyError:
                print('Ошибка из-за неверного прокси, продолжаю работу')
                continue
            else:
                print('Времени прошло - ', datetime.datetime.now() - start)
                get_info(title, url, html)
                break
            i += 1
    print('Закончил работу за - ', datetime.datetime.now() - st)


def get_info(title, url, html):
    """
        Получаем всю инфу о продукте.
    :return:
    """
    print('Достаю инфу с спарсенного сайта.')
    try:
        dict_of_info = {'Картинка': (lambda x: x[x.rfind('https'):-2]) (re.search(r'<meta property=\"og:image\"[^>]*', html).group())}
    except:
        return
    dict_of_info.update({'Ссылка': url})
    key = tuple(map(lambda x: x[x.rfind('>') + 1:], re.findall(r'_2TxqAVjiup[^<]*', html)))
    value = tuple(map(lambda x: x[x.rfind('>') + 1:], re.findall(r'<dd[^<]*', html)))
    for i in range(len(key)):
        dict_of_info.update({key[i]: value[i]})
    write_xlsx(dict_of_info)


def write_xlsx(dict_of_info):
    try:
        wb = openpyxl.load_workbook('2.xlsx')
    except:
        wb = openpyxl.Workbook()
        del wb['Sheet']
    if 'Лист1' not in wb.sheetnames:
        wb.create_sheet('Лист1')
    ws = wb['Лист1']
    row = ws.max_row + 1 if ws.max_row == 1 else ws.max_row + 14
    print('Открыл 2.xlsx.')
    for column, name_of_char in enumerate(dict_of_info.keys()):
        if column == 0:
            with open('1.png', 'wb') as f:
                f.write(requests.get(dict_of_info.get(name_of_char)).content)
            img = Image.open('1.png')
            img.thumbnail((200, 200), Image.ANTIALIAS)
            img.save('1.png')
            img = openpyxl.drawing.image.Image('1.png')
            ws.add_image(img, 'A' + str(row))
        else:
            ws.cell(row=row, column=column + 1).value = name_of_char
        ws.column_dimensions[chr(ord("a") + column)].width = 50
    row = ws.max_row + 1
    del dict_of_info['Картинка']
    for column, name_of_char in enumerate(dict_of_info.values()):
        ws.cell(row=row, column=column + 2).value = name_of_char
    while True:
        try:
            wb.save('2.xlsx')
        except:
            print('Закройте пожалуйста Excel файл 2.xlsx')
            time.sleep(3)
        else:
            wb.close()
            print('Закрыл 2.xlsx.')
            return


if __name__ == '__main__':
    # get_info('sex', 'asd', 'lol')
    read_xlsx()
